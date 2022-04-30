import tkinter as tk
from openpyxl import Workbook
import pyautogui
import time
import keyboard


global res
time.sleep(3)
# Creating a workbook file
wb = Workbook()

# Creating first sheet
ws=wb.active

# Saving workbook with name
wb.save(filename = 'sample.xlsx')


# # Writing Headings of sheet
ws['A1']='S. No.'
ws['B1']='Name'
ws['C1']='Address'
ws['D1']='Phone No.'
ws['E1']='Rating'

wb.save(filename = 'sample.xlsx')

# Part 1- Opening 'view all'
pyautogui.moveTo(671, 130, duration=0.5)
pyautogui.scroll(-650)
pyautogui.moveTo(672, 591, duration=0.5)
pyautogui.click(x=672, y=591, clicks=1, interval=0, button='left')

i=0


# Part 3- Getting data of shops


def pls(i):
    # Part 2- Opening shop 1
    pyautogui.moveTo(81, 355, duration=1)
    time.sleep(4)
    pyautogui.click(x=81, y=355, clicks=1, interval=0, button='left')

    end_block = None
    while end_block==None:

        root = tk.Tk()
        # Name pos = (549,379)
        pyautogui.moveTo(549, 379, duration=1)
        pyautogui.dragTo(1123, 442, duration=0.5)
        pyautogui.hotkey('ctrl', 'c')
        paste = root.clipboard_get()
        ws['B'+str(i)]=paste

        # # Rating pos =(551,482)
        root = tk.Tk()
        pyautogui.click(546, 483, clicks=1, interval=0, button='left')
        rat_pos = pyautogui.locateCenterOnScreen('gogrev.png')
        if rat_pos != None:
            pyautogui.moveTo(rat_pos[0] - 200, rat_pos[1], duration=0.5)
            pyautogui.dragTo(rat_pos[0] - 167, rat_pos[1], duration=0.5)
            pyautogui.hotkey('ctrl', 'c')
            paste = root.clipboard_get()
            if len(paste)<6:
                ws['E' + str(i)] = paste
            else:
                ws['E' + str(i)] = 'Null'

        else:
            root = tk.Tk()
            pyautogui.click(546, 483, clicks=1, interval=0, button='left')
            pyautogui.moveTo(546, 483, duration=1)
            pyautogui.dragTo(582, 483, duration=0.5)
            pyautogui.hotkey('ctrl', 'c')
            paste = root.clipboard_get()
            ws['E'+str(i)]=paste
            if len(paste)<6:
                ws['E' + str(i)] = paste
            else:
                ws['E' + str(i)] = 'Null'


        # # Address pos=(1162,617)
        root = tk.Tk()
        add_pos=pyautogui.locateCenterOnScreen('address.png')
        x_dif = 573
        y_dif = 36
        # (x=1162, y=643)
        if add_pos!=None:
            paste = None
            pyautogui.moveTo(add_pos[0] + 40, add_pos[1], duration=0.5)
            pyautogui.dragTo(add_pos[0] + 573, add_pos[1] + 36, duration=0.5)
            pyautogui.hotkey('ctrl', 'c')
            paste = root.clipboard_get()
            ws['C'+str(i)]=paste
        else:
            ws['C' + str(i)] = 'Null'


        # # Phone pos=(614,681)
        root = tk.Tk()
        phone_pos = pyautogui.locateCenterOnScreen('phone.png')
        if phone_pos != None:
            paste=None
            pyautogui.moveTo(phone_pos[0] + 34, phone_pos[1], duration=0.5)
            pyautogui.dragTo(phone_pos[0] + 200, phone_pos[1], duration=0.5)
            pyautogui.hotkey('ctrl', 'c')
            paste = root.clipboard_get()
            ws['D' + str(i)] = paste
        else:
            ws['D'+str(i)]='Null'

        wb.save(filename='sample.xlsx')


        i=i+1

        # scroll pos= (512,401)
        pyautogui.moveTo(212, 399, duration=0.2)
        pyautogui.scroll(-160)
        pyautogui.click(x=190, y=440, clicks=1, interval=0, button='left')
        pyautogui.click(x=190, y=440, clicks=1, interval=0, button='left')
        if keyboard.is_pressed("q"):
            return pls(i+1)
        else:
            pass


pls(2)

